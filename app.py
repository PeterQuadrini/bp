import requests
from flask import Flask, render_template, session, request, redirect, url_for
from flask_session import Session  # https://pythonhosted.org/Flask-Session
import json
import os
import openpyxl

app = Flask(__name__)
app.secret_key = os.urandom(24)

currentId = None
buddy = None

@app.route("/")
def home():
    if not session.get("user"):
        return render_template('index2.html')
    wb = openpyxl.load_workbook('static/excel.xlsx')
    ws = wb['rel']
    values = [ws.cell(row=i,column=2).value for i in range(1,ws.max_row+1)]
    print('valuesss '+ str(values))
    return render_template('index.html', buddies = values)

@app.route("/pag2")
def index2():
    if not session.get("user"):
        return redirect(url_for('home'))
    return render_template('index.html')

@app.route("/wrong")
def index3():
    return render_template('wrong.html', buddy = buddy)


@app.route('/test', methods=['POST'])
def test():
    data = json.loads(request.data)
    print(data)
    session["user"] = data['idTokenClaims']
    print("okok")
    print(session)
    return '/'

@app.route('/logoutSession', methods=['POST'])
def testLog():
    session.clear()
    print("sessionsssssss")
    print(session)
    return redirect('/')

@app.route('/writeExcel', methods = ['POST'])
def writeExcel():
    ex = openpyxl.load_workbook('static/excel.xlsx')
    exRel = ex['rel']
    print('max rowsssssssS: ' + str(exRel.max_row))
    data = json.loads(request.data)
    new_row = (data['employee'], data['buddy'])
    exRel.append(new_row)
    ex.save('static/excel.xlsx')
    buddy = data['buddy']
    return '/'

@app.route('/checkIfAvailable', methods = ['GET'])
def checkAvailable():
    id = request.args.get('id')
    print(id)
    wb = openpyxl.load_workbook('static/excel.xlsx')
    ws = wb['rel']
    values = [ws.cell(row=i,column=2).value for i in range(1,ws.max_row+1)]
    response = 'available'
    for value in values:
        if value == id:
            print('heyheyheyhtye')
            response = 'busy'
            return 'response', 300
    return response, 200

 
if __name__ == "__main__":
    app.run(host="localhost", debug=True)